import cv2 
from pytesseract import image_to_string
from pdf2image import convert_from_path
import os 
import re
import random
import openpyxl
import shutil
import glob

from pathlib import Path
from datetime import datetime
import shutil

def getName(filename,pdfpath):
    filepath=f"{pdfpath}\{filename}.pdf"
    content = get_text_from_any_pdf(filepath)
    name =""
    # i=1
    # for element in content:
    #     print(i,"=",element)
    #     i=i+1

    # namestartindex=content.find(':')+1
    # # print("namestartindex=",namestartindex)

    # nameendindex=content.find('-')+1
    # print("nameendindex=",nameendindex)

    # lines=content.splitlines()
    i=1
    PATIENT_LINE=""
    # print(content.splitlines(True))
    if content.splitlines(True):
        for line in content.splitlines():
            if "PATIENT" in line:
              PATIENT_LINE=line
            #   print(i,"=",line)
              break
            i=i+1
    if len(PATIENT_LINE)==0:
        return name
    namestartindex=PATIENT_LINE.find(':')+1
    # print("namestartindex=",namestartindex)

    nameendindex=PATIENT_LINE.find('-')
    # print("nameendindex=",nameendindex)
    
    name=PATIENT_LINE[namestartindex:nameendindex].strip()
    
    return name

def get_text_from_any_pdf(pdf_file):
    """
    @desc: this function is our final system combining the previous functions
    
    @params:
        - file: the original PDF File
    
    @returns:
        - the textual content of ALL the pages
    """
    images = convert_pdf_to_img(pdf_file)
    final_text = ""
    i=0
    for pg, img in enumerate(images):
        
        final_text += convert_image_to_text(img)
        #print("Page n°{}".format(pg))
        # print(convert_image_to_text(img))
        i=i+1
  
    return final_text
def getDigitonly(data):
    data=data.strip()
    if data.isnumeric() or data=="" or len(data)==1: 
        # print("data.isnumeric()= ",data.isnumeric() )
        if data.isnumeric() !=True:
            return 0
        return data
    if len(data)==2:
        data1=data[0]
        if data1.isnumeric(): 
            return data1
    if len(data)>=3:
        data1=data[0]
        data2=data[1]
        data3=data[2]
        if data1.isnumeric() and  data2.isnumeric() and data3.isnumeric() : 
             return data1+""+data2+""+data3
        if data1.isnumeric() and  data2.isnumeric() : 
             return data1+""+data2
        if data1.isnumeric() : 
             return data1 
  
    return ""
def convert_image_to_text(file):
    """
    @desc: this function extracts text from image
    
    @params:
        - file: the image file to extract the content
    
    @returns:
        - the textual content of single image
    """
    # img = cv2.imread("work/2.jpg")
    myconfig=r"--psm 11 --oem 3"
    myconfig=r"--oem 3 --psm 6"
    # myconfig='--psm 10 --oem 3 -c tessedit_char_whitelist=abcr0123456789'
    text = image_to_string(file,lang='eng',config=myconfig)
    # text = image_to_string(file)

    return text
def convert_pdf_to_img(pdf_file):
    """
    @desc: this function converts a PDF into Image
    
    @params:
        - pdf_file: the file to be converted
    
    @returns:
        - an interable containing image format of all the pages of the PDF
    """
    return convert_from_path(pdf_file)
def createdir(path):
    current_directory = os.getcwd()
    final_directory = os.path.join(current_directory, path)
    if not os.path.exists(final_directory):
     os.makedirs(final_directory)
    return

def getpoint(frontname,positionname,orgimage,ouputimagepath,x,y,w,h):
    result=0

    box=cv2.rectangle(orgimage.copy(), (x, y), (w, h), (255,0,255), 2)
    cv2.imwrite(f"{ouputimagepath}\ {frontname}c_{positionname}.jpg", box)
    cropped_acrimage = orgimage[y:h, x:w]
    result1=convert_image_to_text(cropped_acrimage).strip()
    # print(f"{frontname}c_{positionname}=",result1)
    result=getDigitonly(result1)
    # print("result==",result)
    if debug:
        print(f"==={frontname}c_{positionname}?{result1}=",result)
    if result =="":
        result=0
    return int(result)

def getdata(frontname,positionname,orgimage,ouputimagepath,x,y,w,h):
    result=0
    box=cv2.rectangle(orgimage.copy(), (x, y), (w, h), (0,0,255), 2)
    cv2.imwrite(f"{ouputimagepath}\ {frontname}c_{positionname}.jpg", box)
    cropped_acrimage = orgimage[y:h, x:w]
    result=convert_image_to_text(cropped_acrimage).strip()
    if debug:
     print(f"{frontname}c_{positionname}=",result)
    # result=getDigitonly(result)
    # print(f"==={frontname}c_{positionname}=",result)
    return result
def wirteXls(index,xrow,name,output4points,fname):
    
    # shutil.copy2('data\Template.xlsx', '/dst/dir/newname.ext')
  
    # fname=f"{xlspath}\Audiometry.xlsx"
    # shutil.copy2('data\Template.xlsx', fname)
    workbook = openpyxl.load_workbook(fname)
    sheet = workbook['Sheet1']
    # change the header name
    sheet.cell(row=xrow, column=1).value = index
    sheet.cell(row=xrow, column=4).value = name
    sheet.cell(row=xrow, column=8).value = output4points[0][0]
    sheet.cell(row=xrow, column=9).value = output4points[0][1]
    sheet.cell(row=xrow, column=10).value = output4points[0][2]
    sheet.cell(row=xrow, column=11).value = output4points[0][3]
    sheet.cell(row=xrow, column=12).value = output4points[0][4]
    sheet.cell(row=xrow, column=13).value = output4points[0][5]


    sheet.cell(row=xrow, column=20).value = output4points[1][0]
    sheet.cell(row=xrow, column=21).value = output4points[1][1]
    sheet.cell(row=xrow, column=22).value = output4points[1][2]
    sheet.cell(row=xrow, column=23).value = output4points[1][3]
    sheet.cell(row=xrow, column=24).value = output4points[1][4]
    sheet.cell(row=xrow, column=25).value = output4points[1][5]

    sheet.cell(row=xrow, column=14).value = output4points[2][0]
    sheet.cell(row=xrow, column=15).value = output4points[2][1]
    sheet.cell(row=xrow, column=16).value = output4points[2][2]
    sheet.cell(row=xrow, column=17).value = output4points[2][3]
    sheet.cell(row=xrow, column=18).value = output4points[2][4]
    sheet.cell(row=xrow, column=19).value = output4points[2][5]

    sheet.cell(row=xrow, column=26).value = output4points[3][0]
    sheet.cell(row=xrow, column=27).value = output4points[3][1]
    sheet.cell(row=xrow, column=28).value = output4points[3][2]
    sheet.cell(row=xrow, column=29).value = output4points[3][3]
    sheet.cell(row=xrow, column=30).value = output4points[3][4]
    sheet.cell(row=xrow, column=31).value = output4points[3][5]
    # save the changes
    workbook.save(fname)
    return
def convert4Points(pdffilename,path):
    outputpoint= [[int(0) for i in range(6)] for j in range(4)]
    # print(outputpoint)
    index=pdffilename+"_"+str(random.randint(10000,99999))
    wpath = f"work\{index}"
    edgepath = f"work\{index}_edge"
    createdir(wpath)
    createdir(edgepath)
# Create the directory 
# 'GeeksForGeeks' in 
# '/home / User / Documents' 

 
    # index = random.randit(11111,99999)
    # index=1
    path_to_pdf = f'{path}\{pdffilename}.pdf'
    pathorgimage = f"{wpath}\org.jpg"
    pathcutrect = f"{wpath}\cuttable.jpg"
    pathcutrect2=f"{wpath}\cuttable2.jpg"
    tcpath=f"{wpath}\thresholdcropped.jpg"
    ipathtest=f"{wpath}\invertcropped.jpg"
    igpathtest=f"{wpath}\invertgraycropped.jpg"
    dpathtest=f"{wpath}\dilate.jpg"
    dgpathtest=f"{wpath}\dilategray.jpg"
    cipathtest=f"{wpath}\contour.jpg"
    acrpath=f"{wpath}\ ac_R.jpg"
    # acr250path=f"{wpath}\ acr250.jpg"
    # acrpointimagepath=f"{wpath}"

    pdfimage=convert_pdf_to_img(path_to_pdf)
    pdfimage[0].save(pathorgimage,'JPEG')
    orgimage = cv2.imread(pathorgimage)
    cropped_image = orgimage[1055:1200, 50:1650]
    cv2.imwrite(pathcutrect, cropped_image)
    img = cv2.imread(pathcutrect)
    grayimage = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)   
    cv2.imwrite(pathcutrect2, grayimage)

    thresholdcropped_image= cv2.threshold(grayimage, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    cv2.imwrite(tcpath, thresholdcropped_image)
    invertcropped_image = cv2.bitwise_not(thresholdcropped_image)
    cv2.imwrite(ipathtest, invertcropped_image)
    imgrectblack1 = cv2.imread(ipathtest)
    ignvertcropped_image = cv2.bitwise_not(grayimage)
    cv2.imwrite(igpathtest, ignvertcropped_image)
    imgrectblack = cv2.imread(igpathtest)
    idilatecropped_image = cv2.dilate(invertcropped_image, None, iterations=5)
    cv2.imwrite(dpathtest, idilatecropped_image)
    igdilatecropped_image = cv2.dilate(ignvertcropped_image, None, iterations=5)
    cv2.imwrite(dgpathtest, igdilatecropped_image)
    # # 
    # edged = cv2.Canny(idilatecropped_image, 30, 200) 
    contours,hierarchy = cv2.findContours(idilatecropped_image, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    image_with_all_contours = grayimage.copy()
    contoursimage=cv2.drawContours(image_with_all_contours,contours, -1, (0, 255, 0), 3)

    cv2.imwrite(cipathtest, contoursimage)

    acrx=0
    acry=0
    acrw=0
    acrh=0
    filerowcount=0

    for i, contour in enumerate(contours):
        area = cv2.contourArea(contour)
        # print("area=",area)
        x, y, w, h = cv2.boundingRect(contour)    
        # print(f"{i}=",cv2.boundingRect(contour))
        approx = cv2.approxPolyDP(contour, 0.009 * cv2.arcLength(contour, True), True) 
        cimage=cv2.drawContours(grayimage, [approx], 0, (0, 0, 255), 5) 
        cropped_image = img[y:y+h, x:x+w]
        cv2.imwrite(f"{edgepath}/{i}.jpg", cropped_image)
        data=convert_image_to_text(cropped_image)
        if area> 10000:
         continue
        if "ACR" in data:       
            acrx=x
            acry=y
            acrw=w
            acrh=h
            # print(f"{i}={acrx},{acry},{acrw},{acrh}=",data)
            break
    # acr image rectangle point
    # grayimage = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 
    grayimage = cv2.cvtColor(imgrectblack, cv2.COLOR_BGR2GRAY)   
    grayimage1 = cv2.cvtColor(imgrectblack1, cv2.COLOR_BGR2GRAY)   
    acrrectx=acrx
    acrrecty=acry
    acrrectw=acrx + acrw
    acrrecth=acry + acrh-11
    box=cv2.rectangle(grayimage.copy(), (acrrectx, acrrecty), (acrrectw, acrrecth), (0,0,255), 2)
    cv2.imwrite(f"{acrpath}", box)

    cropped_acrimage = grayimage[acrrecty:acrrecth, acrrectx:acrrectw]
    cropped_acr_data=convert_image_to_text(cropped_acrimage)
    if debug:
        print("acr_data=",cropped_acr_data)
        
    # acrR 250 data
    acrrectx=acrx+190
    acrrecty=acry-5
    acrrectw=acrrectx + acrw-10
    acrrecth=acry + acrh-24
    # box250=cv2.rectangle(grayimage.copy(), (acrrectx, acrrecty), (acrrectw, acrrecth), (0,0,255), 2)
    # cv2.imwrite(f"{acr250path}", box250)
    # cropped_acrimage = grayimage[acrrecty:acrrecth, acrrectx:acrrectw]
    # acr_250=convert_image_to_text(cropped_acrimage).strip()
    # print("acr_250=",acr_250)
    outputpoint[0][0]=getpoint("a","R_250",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    # print("bcr=",outputpoint[0][0])
    # acrR 500 data
    acrrectx=acrx+235
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-20
    outputpoint[0][1]=getpoint("a","R_500",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # acrR 1k data
    acrrectx=acrx+323
    acrrecty=acry-1
    acrrectw=acrrectx + acrw+2
    acrrecth=acry + acrh-21
    outputpoint[0][2]=getpoint("a","R_1k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    # print("bcr1=",outputpoint[0][2])
    # acrR 2k data
    acrrectx=acrx+415
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24
    outputpoint[0][3]=getpoint("a","R_2k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)


    # acrR 4k data
    acrrectx=acrx+505
    acrrecty=acry-5
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24+6
    outputpoint[0][4]=getpoint("a","R_4k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    # print("bcr=",outputpoint[0][4])
    # acrR 8k data
    acrrectx=acrx+595
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24+2
    outputpoint[0][5]=getpoint("a","R_8k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    print("bcr=",outputpoint[0][5])
    # acrL 250 data
    acrrectx=acrx+1120
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24
    outputpoint[1][0]=getpoint("a","L_250",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # acrL 500 data
    acrrectx=acrx+1165
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24
    outputpoint[1][1]=getpoint("a","L_500",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # acrL 1k data
    acrrectx=acrx+1255
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24
    outputpoint[1][2]=getpoint("a","L_1k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # acrL 2k data
    acrrectx=acrx+1345
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24+2
    outputpoint[1][3]=getpoint("a","L_2k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # acrL 4k data
    acrrectx=acrx+1435
    acrrecty=acry
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-24
    outputpoint[1][4]=getpoint("a","L_4k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # acrL 8k data
    acrrectx=acrx+1525
    acrrecty=acry-5
    acrrectw=acrrectx + acrw
    acrrecth=acry + acrh-22
    outputpoint[1][5]=getpoint("a","L_8k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    print("bcL555=",outputpoint[1][5])


    # identify 4row or 2row data
    acrrectx=acrx
    acrrecty=acry+23
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20
    acrL_data=getdata("b","r",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    buffery=23
    if debug:
        print("bc_L_data1===",acrL_data)
    if "ACR" in acrL_data:
        buffery=46
        
        acrrectx=acrx
        acrrecty=acry+buffery
        acrrectw=acrrectx + acrw
        acrrecth=acrrecty +acrh-20
        acrL_data=getdata("b","R",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    if debug:
        print("bc_L_data2===",acrL_data)

    # bcR 250 data
    acrrectx=acrx+180
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw+20
    acrrecth=acrrecty +acrh-18
    outputpoint[2][0]=getpoint("b","R_250",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    # print("bcr=",outputpoint[2][0])
    # bcR 500 data
    acrrectx=acrx+235
    acrrecty=acry+buffery-5
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20+5
    outputpoint[2][1]=getpoint("b","R_500",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    print("bcr11=",outputpoint[2][1])

    if (outputpoint[2][1] <10 and outputpoint[2][1]!=5) or outputpoint[2][1]==0 :
        acrrecty=acrrecty-3
        acrrecth=acrrecth+3
        outputpoint[2][1]=getpoint("b","R_500",grayimage1.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
        print("bcr111=",outputpoint[2][1])
    # bcR 1k data
    acrrectx=acrx+325
    acrrecty=acry+buffery-2
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-18
    outputpoint[2][2]=getpoint("b","R_1k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    print("bcr115=",outputpoint[2][2])
    if (outputpoint[2][2] <10 and outputpoint[2][2]!=5) or  outputpoint[2][2]==0 or (outputpoint[2][2] %5 !=0 and outputpoint[2][2]>9):
            acrrectx=acrx+325
            acrrecty=acry+buffery
            acrrectw=acrrectx + acrw
            acrrecth=acrrecty +acrh-20
            outputpoint[2][2]=getpoint("b","R_1k",grayimage1.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
            print("bcr1115=",outputpoint[2][2])

    # bcR 2k data
    acrrectx=acrx+415
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20
    outputpoint[2][3]=getpoint("b","R_2k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # bcR 4k data
    acrrectx=acrx+505
    acrrecty=acry+buffery-3
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20
    outputpoint[2][4]=getpoint("b","R_4k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    print("bcr114=",outputpoint[2][4])
    # bcR 8k data
    acrrectx=acrx+595
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-24
    outputpoint[2][5]=getpoint("b","R_8k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)


    # bcL 250 data
    acrrectx=acrx+1120
    acrrecty=acry+buffery-2
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20
    outputpoint[3][0]=getpoint("b","L_250",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    print("bcr300=",outputpoint[3][0])

    if (outputpoint[3][0] <10 and outputpoint[3][0]!=5) or outputpoint[3][0]==0 :
            acrrectx=acrx+1120
            acrrecty=acry+buffery-3
            acrrectw=acrrectx + acrw
            acrrecth=acrrecty +acrh-18
            outputpoint[3][0]=getpoint("b","L_250",grayimage1.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
            print("bcr300=",outputpoint[3][0])

    # bcL 500 dataBG BNUSFD
    
    acrrectx=acrx+1165
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-24
    outputpoint[3][1]=getpoint("b","L_500",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # bcL 1K data
    acrrectx=acrx+1260
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-22
    outputpoint[3][2]=getpoint("b","L_1k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # print("buffery=",buffery)
    # bcL 2K data
    acrrectx=acrx+1345
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20
    outputpoint[3][3]=getpoint("b","L_2k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    if (outputpoint[3][3] <10 and outputpoint[3][3]!=5) or outputpoint[3][3]==0 :
        acrrectx=acrx+1345
        acrrecty=acry+buffery
        acrrectw=acrrectx + acrw
        acrrecth=acrrecty +acrh-20
        outputpoint[3][3]=getpoint("b","L_2k",grayimage1.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)


    # bcL 4K data
    acrrectx=acrx+1435
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20
    outputpoint[3][4]=getpoint("b","L_4k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)

    # bcL 8K data
    acrrectx=acrx+1525
    acrrecty=acry+buffery
    acrrectw=acrrectx + acrw
    acrrecth=acrrecty +acrh-20
    outputpoint[3][5]=getpoint("b","L_8k",grayimage.copy(),wpath,acrrectx,acrrecty,acrrectw,acrrecth)
    
    return outputpoint
def fileWrite(filepath,data):
    d_date = datetime.now()
    reg_format_date = d_date.strftime("%Y-%m-%d %I:%M:%S %p")
    data=f"\n{reg_format_date}>> {data}"

    with open(f"{filepath}", 'a') as file:
        file.write(data)
        print(data)
        # file.close()
    return
def remove(path):
    """ param <path> could either be relative or absolute. """
    if os.path.isfile(path) or os.path.islink(path):
        os.remove(path)  # remove the file
    elif os.path.isdir(path):
        shutil.rmtree(path)  # remove dir and all contains
    
    return
def cleanup(outputpath):

    remove("Completed.log")
    remove("logs")
    remove("work")
    remove(outputpath)
    return
def doConvert(path):
    outputpath=f"{path}\output"
    cleanup(outputpath)
    createdir(outputpath)
    templatefilename="Audiometry.xlsx"
    fname=f"{outputpath}\{templatefilename}"
   
    shutil.copy2(f"data\{templatefilename}", fname)

    statuslogfilepath ="logs"   
    statuslogfile=f"{statuslogfilepath}\status.log"
    createdir(statuslogfilepath)

    
    files=glob.glob(f"{path}\*.pdf")
    fileWrite(statuslogfile,files)
    print(files)
    i=0
    l=len(files)
    row=3
    for file in files:
     i+=1
     basefilename = Path(file).stem
     progressmsg=f"{i}/{l}={basefilename}====================conversion Inprgress"
    #  print(f"{i}/{l}={basefilename}====================conversion Inprgress")
     fileWrite(statuslogfile,progressmsg)
     name=getName(basefilename,"input")
     if name=="":
        progressmsg=f"{i}/{l}={basefilename}======XLSX conversion skipped bacause of invalid pdf report file!"
        fileWrite(statuslogfile,progressmsg)
        continue
         

     
     outputpoint=convert4Points(basefilename,"input")
     wirteXls(i,row,name,outputpoint,fname)


     progressmsg=f"{i}/{l}={basefilename}======XLSX conversion successful!"
     fileWrite(statuslogfile,progressmsg)
    #  print(f"{i}/{l}={basefilename}======XLSX conversion successful!")
     
     row+=1
     if debug:
        print("outputpoint",outputpoint)
        print("name",name)
        print(basefilename)
    progressmsg="================================================================================================================"
    fileWrite(statuslogfile,progressmsg)
    progressmsg="========================================Overall conversion has been completed==================================="
    fileWrite(statuslogfile,progressmsg)
    progressmsg="================================================================================================================"
    fileWrite(statuslogfile,progressmsg)
    fileWrite("Completed.log","completed")
    return 
debug=False

doConvert("input")

# print("================================================================================================================")

# print("========================================Overall conversion has been completed===================================")
# print("================================================================================================================")

# if debug:
    # print("outputpoint",outputpoint)
    # print("name",name)
