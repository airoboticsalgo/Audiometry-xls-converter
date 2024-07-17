import cv2 


import os 
import re
import random
import openpyxl
import shutil
import glob
import lib.fileoperation as fileoperation
from pathlib import Path
from datetime import datetime
import shutil
import lib.pdfimagetextconversion as pdfimagetextconversion
import lib.puretoneaveragetable as puretoneaveragetable
import lib.personaldata as personaldata







def wirteXls(index,xrow,p_data,output4points,outputpoint2,fname):
    
    # shutil.copy2('data\Template.xlsx', '/dst/dir/newname.ext')
  
    # fname=f"{xlspath}\Audiometry.xlsx"
    # shutil.copy2('data\Template.xlsx', fname)
    workbook = openpyxl.load_workbook(fname)
    sheet = workbook['Sheet1']
    # change the header name
    dat=""
    
    sheet.cell(row=xrow, column=1).value = index
    if p_data[1]!=None and p_data[1].isnumeric():
        sheet.cell(row=xrow, column=2).value = int(p_data[1])
    sheet.cell(row=xrow, column=3).value = p_data[0]
    if p_data[2]!=None and p_data[2].isnumeric():
        sheet.cell(row=xrow, column=4).value = int(p_data[2])
    sheet.cell(row=xrow, column=5).value = p_data[3]
    sheet.cell(row=xrow, column=6).value = p_data[4]

    sheet.cell(row=xrow, column=7).value = output4points[0][0]
    sheet.cell(row=xrow, column=8).value = output4points[0][1]
    sheet.cell(row=xrow, column=9).value = output4points[0][2]
    sheet.cell(row=xrow, column=10).value = output4points[0][3]
    sheet.cell(row=xrow, column=11).value = output4points[0][4]
    sheet.cell(row=xrow, column=12).value = output4points[0][5]

    sheet.cell(row=xrow, column=13).value = output4points[4][0]
    sheet.cell(row=xrow, column=14).value = output4points[4][1]
    sheet.cell(row=xrow, column=15).value = output4points[4][2]
    sheet.cell(row=xrow, column=16).value = output4points[4][3]
    sheet.cell(row=xrow, column=17).value = output4points[4][4]
    sheet.cell(row=xrow, column=18).value = output4points[4][5]

    sheet.cell(row=xrow, column=19).value = output4points[2][0]
    sheet.cell(row=xrow, column=20).value = output4points[2][1]
    sheet.cell(row=xrow, column=21).value = output4points[2][2]
    sheet.cell(row=xrow, column=22).value = output4points[2][3]
    sheet.cell(row=xrow, column=23).value = output4points[2][4]
    sheet.cell(row=xrow, column=24).value = output4points[2][5]

    sheet.cell(row=xrow, column=25).value = output4points[6][0]
    sheet.cell(row=xrow, column=26).value = output4points[6][1]
    sheet.cell(row=xrow, column=27).value = output4points[6][2]
    sheet.cell(row=xrow, column=28).value = output4points[6][3]
    sheet.cell(row=xrow, column=29).value = output4points[6][4]
    sheet.cell(row=xrow, column=30).value = output4points[6][5]

    sheet.cell(row=xrow, column=31).value = outputpoint2[0][0]
    sheet.cell(row=xrow, column=32).value = outputpoint2[0][2]


    sheet.cell(row=xrow, column=33).value = output4points[1][0]
    sheet.cell(row=xrow, column=34).value = output4points[1][1]
    sheet.cell(row=xrow, column=35).value = output4points[1][2]
    sheet.cell(row=xrow, column=36).value = output4points[1][3]
    sheet.cell(row=xrow, column=37).value = output4points[1][4]
    sheet.cell(row=xrow, column=38).value = output4points[1][5]

    sheet.cell(row=xrow, column=39).value = output4points[5][0]
    sheet.cell(row=xrow, column=40).value = output4points[5][1]
    sheet.cell(row=xrow, column=41).value = output4points[5][2]
    sheet.cell(row=xrow, column=42).value = output4points[5][3]
    sheet.cell(row=xrow, column=43).value = output4points[5][4]
    sheet.cell(row=xrow, column=44).value = output4points[5][5]


    sheet.cell(row=xrow, column=45).value = output4points[3][0]
    sheet.cell(row=xrow, column=46).value = output4points[3][1]
    sheet.cell(row=xrow, column=47).value = output4points[3][2]
    sheet.cell(row=xrow, column=48).value = output4points[3][3]
    sheet.cell(row=xrow, column=49).value = output4points[3][4]
    sheet.cell(row=xrow, column=50).value = output4points[3][5]

    sheet.cell(row=xrow, column=51).value = output4points[7][0]
    sheet.cell(row=xrow, column=52).value = output4points[7][1]
    sheet.cell(row=xrow, column=53).value = output4points[7][2]
    sheet.cell(row=xrow, column=54).value = output4points[7][3]
    sheet.cell(row=xrow, column=55).value = output4points[7][4]
    sheet.cell(row=xrow, column=56).value = output4points[7][5]

    sheet.cell(row=xrow, column=57).value = outputpoint2[0][1]
    sheet.cell(row=xrow, column=58).value = outputpoint2[0][3]


    # save the changes
    workbook.save(fname)
    return
def convert4Points(pdffilename,path):
    outputpoint= [[int(0) for i in range(6)] for j in range(8)]
    outputpoint2= [[float(0) for i in range(4)] for j in range(1)]
    # print(outputpoint)
    index=pdffilename+"_"+str(random.randint(100000,999999))
    wpath = f"work\{index}"
    # edgepath = f"work\{index}_edge"
    fileoperation.createdir(wpath)
    # fileoperation.createdir(edgepath)
# Create the directory 
# 'GeeksForGeeks' in 
# '/home / User / Documents' 

 
    # index = random.randit(11111,99999)
    # index=1
    path_to_pdf = f'{path}\{pdffilename}.pdf'
    pathorgimage = f"{wpath}\org.jpg"
    pathcutrect = f"{wpath}\cuttable.jpg"
    pathcutrect2=f"{wpath}\cuttable2.jpg"
    tcpath=f"{wpath}\thresholdcropped.jpg"
    ipathtest=f"{wpath}\invertcropped.jpg"
    igpathtest=f"{wpath}\invertgraycropped.jpg"
    dpathtest=f"{wpath}\dilate.jpg"
    dgpathtest=f"{wpath}\dilategray.jpg"
    cipathtest=f"{wpath}\contour.jpg"
    acrpath=f"{wpath}\ ac_R.jpg"
    avgtonerecttablepath=f"{wpath}\ avgtonerecttable.jpg"
    iavgtonerecttablepath=f"{wpath}\ iavgtonerecttable.jpg"
    igavgtonerecttablepath=f"{wpath}\ igavgtonerecttable.jpg"
    ciavgtonerecttablepath=f"{wpath}\ ciavgtonerecttable.jpg"
    # acr250path=f"{wpath}\ acr250.jpg"
    # acrpointimagepath=f"{wpath}"

    pdfimage=pdfimagetextconversion.convert_pdf_to_img(path_to_pdf)
    pdfimage[0].save(pathorgimage,'JPEG')
    orgimage = cv2.imread(pathorgimage)
    cropped_image = orgimage[1055:1200, 50:1650]
    cv2.imwrite(pathcutrect, cropped_image)
    img = cv2.imread(pathcutrect)
    grayimage = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)   
    # cv2.imwrite(pathcutrect2, grayimage)

    thresholdcropped_image= cv2.threshold(grayimage, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    # cv2.imwrite(tcpath, thresholdcropped_image)
    invertcropped_image = cv2.bitwise_not(thresholdcropped_image)
    cv2.imwrite(ipathtest, invertcropped_image)
    invertimgrectblack = cv2.imread(ipathtest)
    ignvertcropped_image = cv2.bitwise_not(grayimage)
    cv2.imwrite(igpathtest, ignvertcropped_image)
    igninvertimgrectblack = cv2.imread(igpathtest)
    idilatecropped_image = cv2.dilate(invertcropped_image, None, iterations=5)
    # cv2.imwrite(dpathtest, idilatecropped_image)
    # igdilatecropped_image = cv2.dilate(ignvertcropped_image, None, iterations=5)
    # cv2.imwrite(dgpathtest, igdilatecropped_image)
    # # 
    # edged = cv2.Canny(idilatecropped_image, 30, 200) 
    contours,hierarchy = cv2.findContours(idilatecropped_image, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    # image_with_all_contours = grayimage.copy()
    # contoursimage=cv2.drawContours(image_with_all_contours,contours, -1, (0, 255, 0), 3)

    # cv2.imwrite(cipathtest, contoursimage)

    acrx=0
    acry=0
    acrw=0
    acrh=0
    filerowcount=0

    for i, contour in enumerate(contours):
        area = cv2.contourArea(contour)
        # print("area=",area)
        x, y, w, h = cv2.boundingRect(contour)    
        # print(f"{i}=",cv2.boundingRect(contour))
        # approx = cv2.approxPolyDP(contour, 0.009 * cv2.arcLength(contour, True), True) 
        # cimage=cv2.drawContours(grayimage, [approx], 0, (0, 0, 255), 5) 
        cropped_image = img[y:y+h, x:x+w]
        # cv2.imwrite(f"{edgepath}/{i}.jpg", cropped_image)
        data=pdfimagetextconversion.convert_image_to_text(cropped_image)
        if area> 10000:
         continue
        if "ACR" in data:       
            acrx=x
            acry=y
            acrw=w
            acrh=h
            # print(f"{i}={acrx},{acry},{acrw},{acrh}=",data)
            break
    # acr image rectangle point
    # grayimage = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

    grayimage1 = cv2.cvtColor(igninvertimgrectblack, cv2.COLOR_BGR2GRAY)   
    grayimage2 = cv2.cvtColor(invertimgrectblack, cv2.COLOR_BGR2GRAY)   
    # acrrectx=acrx
    # acrrecty=acry
    # acrrectw=acrx + acrw
    # acrrecth=acry + acrh-11
    # box=cv2.rectangle(grayimage.copy(), (acrrectx, acrrecty), (acrrectw, acrrecth), (255,0,255), 2)
    # cv2.imwrite(f"{acrpath}", box)
    

    # cropped_acrimage = grayimage[acrrecty:acrrecth, acrrectx:acrrectw]
    # cropped_acr_data=pdfimagetextconversion.convert_image_to_text(cropped_acrimage)
    # print("acr_data=",cropped_acr_data)

    # if debug:
    #     print("acr_data=",cropped_acr_data)
        
    isidentify4rowstable,outputpoint=pdfimagetextconversion.getdata_AC_BC_R_L_Thr_4row(grayimage1,grayimage2,wpath,acrx,acry,acrw,acrh,outputpoint)
   
    # print("isidentify4rowstable1==",isidentify4rowstable)
    if isidentify4rowstable:
        outputpoint=pdfimagetextconversion.getdata_AC_BC_R_L_mask_4row(grayimage1,grayimage2,wpath,acrx,acry,acrw,acrh,outputpoint)
    # else:
    #     outputpoint= [[int(0) for i in range(6)] for j in range(8)]


    # get avg data
    outputpoint2=puretoneaveragetable.getavgdata(img=orgimage,outputpoint=outputpoint2,wpath=wpath,avgtonerecttablepath=avgtonerecttablepath,iavgtonerecttablepath=iavgtonerecttablepath,igavgtonerecttablepath=igavgtonerecttablepath,ciavgtonerecttablepath=ciavgtonerecttablepath)


    return outputpoint,outputpoint2


def fileWrite(filepath,data):
    d_date = datetime.now()
    reg_format_date = d_date.strftime("%Y-%m-%d %I:%M:%S %p")
    data=f"\n{reg_format_date}>> {data}"

    with open(f"{filepath}", 'a') as file:
        file.write(data)
        print(data)
        # file.close()
    return
def remove(path):
    """ param <path> could either be relative or absolute. """
    if os.path.isfile(path) or os.path.islink(path):
        os.remove(path)  # remove the file
    elif os.path.isdir(path):
        shutil.rmtree(path)  # remove dir and all contains
    
    return
def cleanup(outputpath):

    remove("Completed.log")
    remove("logs")
    remove("work")
    remove(outputpath)
    return
def doConvert(path):
    outputpath=f"{path}\output"
    cleanup(outputpath)
    fileoperation.createdir(outputpath)
    templatefilename="Audiometry.xlsx"
    fname=f"{outputpath}\{templatefilename}"
   
    shutil.copy2(f"data\{templatefilename}", fname)

    statuslogfilepath ="logs"   
    statuslogfile=f"{statuslogfilepath}\status.log"
    fileoperation.createdir(statuslogfilepath)

    
    files=glob.glob(f"{path}\*.pdf")
    fileWrite(statuslogfile,files)
    print(files)
    i=0
    l=len(files)
    row=3
    for file in files:
     i+=1
     basefilename = Path(file).stem
     progressmsg=f"{i}/{l}={basefilename}====================conversion Inprgress"
    #  print(f"{i}/{l}={basefilename}====================conversion Inprgress")
     fileWrite(statuslogfile,progressmsg)
     p_data=personaldata.getpersonaldata(basefilename,"input")
    #  name=
    
     if p_data[0]=="":
        progressmsg=f"{i}/{l}={basefilename}======XLSX conversion skipped bacause of invalid pdf report file!"
        fileWrite(statuslogfile,progressmsg)
        continue
         

    #  name=p_data[0]
     outputpoint,outputpoint2=convert4Points(basefilename,"input")
     wirteXls(i,row,p_data,outputpoint,outputpoint2,fname)


     progressmsg=f"{i}/{l}={basefilename}======XLSX conversion successful!"
     fileWrite(statuslogfile,progressmsg)
    #  print(f"{i}/{l}={basefilename}======XLSX conversion successful!")
     
     row+=1
    #  if debug:
    #     print("outputpoint",outputpoint)
    #     print("name",name)
    #     print(basefilename)
    progressmsg="================================================================================================================"
    fileWrite(statuslogfile,progressmsg)
    progressmsg="========================================Overall conversion has been completed==================================="
    fileWrite(statuslogfile,progressmsg)
    progressmsg="================================================================================================================"
    fileWrite(statuslogfile,progressmsg)
    fileWrite("Completed.log","completed")
    return 
# debug=False

# doConvert("input")

# print("================================================================================================================")

# print("========================================Overall conversion has been completed===================================")
# print("================================================================================================================")

# if debug:
    # print("outputpoint",outputpoint)
    # print("name",name)
